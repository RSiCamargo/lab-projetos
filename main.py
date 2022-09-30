# pip install mysql-connector-python
# pip install yagmail
# pip install pixqrcode
# pip install pillow
# pip install openpyxl
# pip install flask-bootstrap

from xml.etree.ElementTree import tostring
from flask import Flask, render_template, request
from flask_bootstrap import Bootstrap
from pixqrcode import PixQrCode
from openpyxl import load_workbook
from mysql.connector import Error
import yagmail
import mysql.connector
import datetime
import os

username = "Usuario Teste"

# region Classes


class User:
    def __init__(self, name, email, pix):
        self.name = name
        self.email = email
        self.pix = pix


class Client:
    def __init__(self, name, email, address, date):
        self.name = name
        self.email = email
        self.address = address
        self.date = date


class Product:
    def __init__(self, name, amount, unitPrice):
        self.name = name
        self.amount = amount
        self.unitPrice = unitPrice

# endregion


# region Banco
# --------------- Cria a conexão com o banco de dados ---------------
def create_connection(host_name, user_name, user_password, db_name):
    connection = None
    try:
        connection = mysql.connector.connect(
            host=host_name,
            user=user_name,
            passwd=user_password,
            database=db_name
        )
        print("Connection to MySQL DB successful")
    except Error as e:
        print(f"The error '{e}' occurred")

    return connection


# --------------- Função para execução de single queries ---------------
def execute_query(connection, query):
    cursor = connection.cursor()
    try:
        cursor.execute(query)
        connection.commit()
        return True
    except Error as e:
        return False

# endregion


# region Consumo de Arquivos
# --------------- Arquivos de Dados ---------------
def readDataFile(file):
    userEmail = file["D3"].value
    userPix = file["D5"].value

    for value in file.iter_rows(min_row=8, min_col=3, max_col=7, values_only=True):
        if (value[0] != None):
            updateUser(value[0], value[2], value[3],
                       value[4], username, userPix, userEmail)
        else:
            break


# --------------- Arquivo de Consumo ---------------
def readExpenseFile(file):
    date = file["D1"].value

    for value in file.iter_rows(min_row=4, min_col=3, max_col=6, values_only=True):
        if (value[0] != None):
            checkClientAndSave(value[0], value[1], value[2], value[3], date)
        else:
            break


# --------------- Arquivo de Estoque ---------------
def readStockFile(file):
    for value in file.iter_rows(min_row=4, min_col=3, max_col=6, values_only=True):
        if (value[0] != None):
            updateStockAndSave(value[0], value[1], id, value[3], value[2])
        else:
            break

# endregion


# region Metodos
def updateUser(clientName, billingDay, email, city, userName, userPix, userEmail):
    connection = create_connection(
        "us-cdbr-east-06.cleardb.net", "b090112be85288", "2f84fdce", "heroku_7324e25c80c3d90")

    print(f"O User {userName} com o Pix {userPix}, adicionou o cliente {clientName}, residente da cidade {city} e com email {email}, com faturamento agendado para o dia {billingDay}")

    cursor = connection.cursor()
    stmt = ("INSERT INTO data (clientName, billingDay, email, city, userName ,userPix, userEmail) VALUES (%s, %s, %s, %s, %s, %s, %s)")
    values = (clientName, billingDay, email,
              city, userName, userPix, userEmail)
    cursor.execute(stmt, values)
    connection.commit()
    cursor.close


def checkClientAndSave(nome, insumo, quantidade, preco, created_at):

    connection = create_connection(
        "us-cdbr-east-06.cleardb.net", "b090112be85288", "2f84fdce", "heroku_7324e25c80c3d90")
    cursor = connection.cursor()

    # Checar se o cliente existe, e salvar no banco (alertar se o cliente não existir)
    query = ("SELECT CASE WHEN EXISTS (SELECT  clientName FROM data  WHERE clientName = %s) THEN 1 ELSE 0 END ")
    name = (nome,)
    cursor.execute(query, name)
    for row in cursor:
        checkName = row[0]

    if (checkName == 1):
        print("Existe no banco")
        print(
            f"O cliente {nome} comprou {insumo}, quantidade {quantidade} e com porcentagem de  {preco},  criado em {created_at}")
        stmt = 'INSERT INTO cons (clientName, product, amount, price ,created_at) VALUES (%s, %s, %s, %s, %s)'
        data = (nome, insumo, quantidade, preco, created_at)
        cursor.execute(stmt, data)
        connection.commit()

        stmt = (
            "SELECT id,amount,product FROM cons WHERE id=(SELECT LAST_INSERT_ID()) ")
        cursor.execute(stmt)
        for row in cursor:
            id = row[0]
            quantidade = row[1]
            insumo = row[2]

        updateStockAndSave(insumo, quantidade, id, "proxUpdate")
        cursor.close
    else:
        print("Cliente não cadastrado no banco")


def updateStockAndSave(insumo, quantidade, id, alerta, preco=0):
    connection = create_connection(
        "us-cdbr-east-06.cleardb.net", "b090112be85288", "2f84fdce", "heroku_7324e25c80c3d90")
    cursor = connection.cursor()
    if (preco != 0):
        # Entrará aqui se o arquivo for para atualizacao de estoque.
        stmt = (
            "INSERT INTO stock (product, amount, price, warning) VALUES (%s, %s , %s, %s)")
        values = (insumo, quantidade, preco, alerta)
        cursor.execute(stmt, values)
        connection.commit()
        cursor.close
        print(
            f"Foi adicionado {quantidade} unidades do insumo {insumo} ajustado ao custo base de {preco} - alerta de baixo estoque settado em {alerta} unidades")
    else:
        # Entrará aqui se o arquivo for para a retirada de estoque.
        stmt = ("UPDATE stock SET amount = amount - %s WHERE product = %s")
        data = (quantidade, insumo)
        cursor.execute(stmt, data)
        connection.commit()
        cursor.close
        print(
            f"Foi retirado {quantidade} unidades do insumo {insumo} no estoque")


def cleanDatabase():
    connection = create_connection(
        "us-cdbr-east-06.cleardb.net", "b090112be85288", "2f84fdce", "heroku_7324e25c80c3d90")
    cursor = connection.cursor()
    stmt = ("DELETE FROM cons WHERE clientName='Cliente Apresentacao'")
    cursor.execute(stmt)
    connection.commit()
    stmt = ("DELETE FROM data WHERE clientName='Cliente Apresentacao';")
    cursor.execute(stmt)
    connection.commit()
    stmt = ("DELETE FROM stock WHERE product = 'Produto 1';")
    cursor.execute(stmt)
    connection.commit()
    stmt = ("DELETE FROM stock WHERE product = 'Produto 2';")
    cursor.execute(stmt)
    connection.commit()
    print("Remoções realizadas no banco")

# endregion


# region Rotinas
def createBilling():
    connection = create_connection(
        "us-cdbr-east-06.cleardb.net", "b090112be85288", "2f84fdce", "heroku_7324e25c80c3d90")
    cursor = connection.cursor()

    # Vai checar quais os clientes finais possuem faturamentos no dia
    print("Gerando Faturamento")
    stmt = ("SELECT CURDATE()")
    cursor.execute(stmt)

    for row in cursor:
        date = row[0]

    day = date.strftime("%d")
    stmt = ("SELECT userPix,cons.clientName,email,city,userName,GROUP_CONCAT(price),GROUP_CONCAT(product),GROUP_CONCAT(amount) FROM data,cons WHERE billingDay = %s and data.clientName = cons.clientName group by userPix")
    d = (day,)
    cursor.execute(stmt, d)
    row = cursor.fetchall()

    for all in row:
        # print(all[0], all[1], all[2], all[3], all[4], all[5], all[6], all[7])
        priceList = all[5].split(',')
        total = sum(float(i) for i in priceList)
        total = int(total * 100)

        currentUser = User(all[4], "null", all[0])
        currentClient = Client(all[1], all[2], all[3], "null")
        currentProduct = Product(all[6].split(
            ','), all[7].split(','), priceList)

        pix = generateQRCode(all[0], all[3], all[4], str(total), all[1])

        qr = f"./static/generatedpix/pix-{all[1]}.png"

        sendBillingMail(qr, currentClient, currentUser, currentProduct, total)

    print("Fechando Faturamento")


def generateQRCode(userPix, city, username, value, clientName):
    pix = PixQrCode(username, userPix, city, value)

    if pix.is_valid():
        pix.save_qrcode(
            "./static/generatedpix", f"pix-{clientName}")
        return pix.generate_code()
    else:
        return "Dados para Pix inválidos."


def checkStock():
    # Vai checar, para cada user, como estao os estoques
    connection = create_connection(
        "us-cdbr-east-06.cleardb.net", "b090112be85288", "2f84fdce", "heroku_7324e25c80c3d90")
    cursor = connection.cursor()
    stmt = ("SELECT GROUP_CONCAT(price),GROUP_CONCAT(product),GROUP_CONCAT(amount),userEmail,userName FROM stock,data WHERE amount <= warning GROUP BY product,amount,userEmail")
    cursor.execute(stmt)
    row = cursor.fetchall()
    for all in row:
        currentUser = User(all[4], all[3], "pix")
        currentProduct = Product(all[1].split(
            ','), all[2].split(','), all[0].split(','))
        sendStockAlertMail(currentProduct, currentUser)
    cursor.close

# endregion


# region Envio Email
# --------------- Envio de Email ---------------
def sendStockAlertMail(product, user):
    app_email = 'labprojetospuc@gmail.com'
    app_password = 'npfyvwfaljvlvplv'  # Token for gmail
    to = user.email

    productList = product.name
    amountList = product.amount
    priceList = product.unitPrice

    list = ""

    for num in range(len(productList)):
        lista = list + \
            f'<span style="border-bottom: 1px solid lightgray;display: inline-block;width: 100%;"><p style="text-align: left;display: inline-block;width: 70%;">{productList[num]}</p><i>{amountList[num]}</i><a style="text-align: right;display: inline-block;width: 15%;position: relative;">R${priceList[num]}</a></span>'

    subject = 'Alerta de produtos esgotando do estoque'
    content = ['<div style="text-align: center;background-color: #e1e1e1;"><div>',
               f'<h2>Olá {user.name},</h2><p style="font-size: large">Este é um email de aviso referente aos produtos que estão acabando de seu estoque. <br> Para adicionar mais no sistema, utilize a mesma template de estoque APENAS com os produtos que irá acrescentar a ele.</p></div>',
               f'<div style="border: solid 1px gray; width: 1000px; margin: 0 auto; background-color: white; box-shadow: 2px 2px #86868621"><header>',
               yagmail.inline("./static/img/title.png"),
               f'</header><section><article><h1>Estoque abaixo do limite</h1><div style="width: 100%;"><a style="width: 65%; padding-left: 3%;display: inline-block;position: relative;font-weight: bold;text-align: left;">Produto</a><a style="width: 12.5%; padding-left: 2.5%;display: inline-block;position: relative;font-weight: bold;text-align: left;">Quantidade</a><a style="width: 9%; padding-left: 1%;display: inline-block;position: relative;font-weight: bold;text-align: left;">Preço(un.)</a></div>',
               lista,
               f'</article></section></div><footer style="margin-top: 10px;color: rgb(143, 143, 143);font-family: sans-serif;"><p style="padding: 0px;margin: 5px;">Ward - Automated Billing Service</p><p style="padding: 0px;margin: 5px;">Av. Padre Cletus Francis Cox, 1661 - Country Club, Poços de Caldas - MG, 37714-620</p><p style="padding: 0px;margin: 5px;">Nos contate em labprojetospuc@gmail.com</p></footer></div>']

    with yagmail.SMTP(app_email, app_password) as yag:
        yag.send(to, subject, content)
        print('Sent email successfully')


def sendBillingMail(qr, client, user, product, total):
    userEmail = 'labprojetospuc@gmail.com'
    app_password = 'npfyvwfaljvlvplv'  # Token for gmail
    to = client.email

    productList = product.name
    amountList = product.amount
    priceList = product.unitPrice
    total = str(total)[:-2]
    virgula = ',00'
    total = total + virgula

    list = ""
    lista = []

    for num in range(len(productList)):
        list = f'<span style="border-bottom: 1px solid lightgray;display: inline-block;width: 100%;"><p style="text-align: left;display: inline-block;width: 70%;">{productList[num]}</p><i>{amountList[num]}</i><a style="text-align: right;display: inline-block;width: 15%;position: relative;">R${priceList[num]}</a></span>'
        lista.append(list)

    items = ["&emsp;{}".format(s) for s in lista]
    items = "".join(items)
    subject = f'Cobrança à {user.name}'
    content = ['<div style="text-align: center;background-color: #e1e1e1;"><div>',
               f'<h2>Olá {client.name},</h2><p style="font-size: large">Este é um email referente à cobrança gerada por {user.name}. Em anexo segue o código pix referente ao valor de R${total}.</p></div>',
               f'<div style="border: solid 1px gray; width: 1000px; margin: 0 auto; background-color: white; box-shadow: 2px 2px #86868621"><header>',
               yagmail.inline("./static/img/title.png"),
               f'</header><section><article><h1>Resumo da compra</h1><div style="width: 100%;"><a style="width: 65%; padding-left: 3%;display: inline-block;position: relative;font-weight: bold;text-align: left;">Produto</a><a style="width: 12.5%; padding-left: 2.5%;display: inline-block;position: relative;font-weight: bold;text-align: left;">Quantidade</a><a style="width: 9%; padding-left: 1%;display: inline-block;position: relative;font-weight: bold;text-align: left;">Preço</a></div>',
               items,
               f'<div style="border-top: 2px solid black"><p style="width: 95%;text-align: right;font-family: sans-serif;font-size: large;font-weight: bold;">Total: R${total}</p></div></article></section></div>',
               yagmail.inline(
                   f"./static/generatedpix/pix-{client.name}.png"),
               '<footer style="margin-top: 10px;color: rgb(143, 143, 143);font-family: sans-serif;"><p style="padding: 0px;margin: 5px;">Ward - Automated Billing Service</p><p style="padding: 0px;margin: 5px;">Av. Padre Cletus Francis Cox, 1661 - Country Club, Poços de Caldas - MG, 37714-620</p><p style="padding: 0px;margin: 5px;">Nos contate em labprojetospuc@gmail.com</p></footer></div>']
    attachments = [qr]

    with yagmail.SMTP(userEmail, app_password) as yag:
        yag.send(to, subject, content, attachments)
        print('Sent email successfully')

# endregion


# ---------------  Running Flask ---------------
# Futuro: Pegar novo do usuario pela sessao/cadastro/etc... e adicionar ou novo do arquivo de upload
app = Flask(__name__)
app._static_folder = '../content'
app.config['UPLOAD_FOLDER'] = "static/uploadedexcel/"


@app.route("/")
def home():
    return render_template('index.html', username=username)


@app.route('/upload', methods=['POST', 'GET'])
def upload():
    if request.method == 'POST':
        f = request.files['file']
        f.save(app.config['UPLOAD_FOLDER'] + f.filename)

        try:
            fn = "static/uploadedexcel/" + os.path.basename(f.filename)
            input_sheet = load_workbook(fn)
            sheet = input_sheet.active
        except ValueError:
            print("Erro no upload de arquivo. Será aceita apenas a extensão .xlsx")

        code = sheet["AA1"].value
        if code == 'expense':
            readExpenseFile(sheet)
            return "Upload do arquivo de consumo efetuado com sucesso!"
        elif code == 'stock':
            readStockFile(sheet)
            return "Upload do arquivo de estoque efetuado com sucesso!"
        elif code == 'dados':
            readDataFile(sheet)
            return "Upload do arquivo de dados efetuado com sucesso!"
        else:
            return 'O arquivo não pode ser validado. Por favor utilize os templates disponíveis no github!'


@app.route('/daily', methods=['POST', 'GET'])
def daily():
    if request.method == 'POST':
        if request.form['help'] == 'Call Routines':
            try:
                createBilling()
            except ValueError:
                return "Erro ao gerar faturamento!"
            try:
                checkStock()
            except ValueError:
                return "Erro ao realizar a verificação do estoque!"

            return "Faturamento e checagem de estoque rodados com sucesso!!"
        elif request.form['help'] == 'Clean DB':
            cleanDatabase()
            return "Dados de teste excluidos do DB"


if __name__ == "__main__":
    app.run()
