# !pip install mysql-connector-python
# !pip install yagmail
# !pip install pixqrcode
# !pip install flask-ngrok
# !pip install pillow

from xml.etree.ElementTree import tostring
from flask_ngrok import run_with_ngrok
from flask import Flask, render_template
from pixqrcode import PixQrCode
from openpyxl import load_workbook
from mysql.connector import Error
import yagmail
import mysql.connector
import datetime
import cgi
import os
import cgitb
cgitb.enable()


class User:
    def __init__(self, name, email, pix):
        self.name = name
        self.email = email
        self.pix = pix


class Client:
    def __init__(self, name, email, address, date):
        self.name = name
        self.email = email
        self.address = address
        self.date = date


class Product:
    def __init__(self, name, amount, unitPrice):
        self.name = name
        self.amount = amount
        self.unitPrice = unitPrice


# --------------- Cria a conexão com o banco de dados ---------------

def create_connection(host_name, user_name, user_password, db_name):
    connection = None
    try:
        connection = mysql.connector.connect(
            host=host_name,
            user=user_name,
            passwd=user_password,
            database=db_name
        )
        print("Connection to MySQL DB successful")
    except Error as e:
        print(f"The error '{e}' occurred")

    return connection

# --------------- Função para execução de single queries ---------------


def execute_query(connection, query):
    cursor = connection.cursor()
    try:
        cursor.execute(query)
        connection.commit()
        return True
    except Error as e:
        return False


# --------------- Arquivo de Dados ---------------
def readDataFile(file):
    username = file["D3"].value
    userEmail = file["D4"].value
    userPix = file["D6"].value

    for value in file.iter_rows(min_row=9, min_col=3, max_col=7, values_only=True):
        if (value[0] != None):
            updateUser(value[0], value[2], value[3],
                       value[4], username, userPix, userEmail)
        else:
            break


def updateUser(clientName, billingDay, email, city, userName, userPix, userEmail):
    connection = create_connection(
        "us-cdbr-east-06.cleardb.net", "b090112be85288", "2f84fdce", "heroku_7324e25c80c3d90")
    # Salvar no banco as infos do usuário, e fazer o update da lista de clientes (trabalhar com tudo minusculo)
    print(f"O User {userName} com o Pix {userPix}, adicionou o cliente {clientName}, residente da cidade {city} e com email {email}, com faturamento agendado para o dia {billingDay}")
    cursor = connection.cursor()
    stmt = ("INSERT INTO data (clientName, billingDay, email, city, userName ,userPix, userEmail) VALUES (%s, %s, %s, %s, %s, %s, %s)")
    values = (clientName, billingDay, email,
              city, userName, userPix, userEmail)
    cursor.execute(stmt, values)
    connection.commit()
    cursor.close


# --------------- Arquivo de Consumo ---------------

def readExpenseFile(file):
    date = file["D1"].value

    for value in file.iter_rows(min_row=4, min_col=3, max_col=6, values_only=True):
        if (value[0] != None):
            checkClientAndSave(value[0], value[1], value[2], value[3], date)
        else:
            break


def checkClientAndSave(nome, insumo, quantidade, preco, created_at):

    connection = create_connection(
        "us-cdbr-east-06.cleardb.net", "b090112be85288", "2f84fdce", "heroku_7324e25c80c3d90")
    cursor = connection.cursor()

    # Checar se o cliente existe, e salvar no banco (alertar se o cliente não existir)
    query = ("SELECT CASE WHEN EXISTS (SELECT  clientName FROM data  WHERE clientName = %s) THEN 1 ELSE 0 END ")
    name = (nome,)
    cursor.execute(query, name)
    for row in cursor:
        checkName = row[0]

    if (checkName == 1):
        print("Existe no banco")
        print(
            f"O cliente {nome} comprou {insumo}, quantidade {quantidade} e com porcentagem de  {preco},  criado em {created_at}")
        stmt = 'INSERT INTO cons (clientName, product, amount, price ,created_at) VALUES (%s, %s, %s, %s, %s)'
        data = (nome, insumo, quantidade, preco, created_at)
        cursor.execute(stmt, data)
        connection.commit()

        stmt = (
            "SELECT id,amount,product FROM cons WHERE id=(SELECT LAST_INSERT_ID()) ")
        cursor.execute(stmt)
        for row in cursor:
            id = row[0]
            quantidade = row[1]
            insumo = row[2]

        updateStockAndSave(insumo, quantidade, id, "proxUpdate")
        cursor.close
    else:
        print("Cliente não cadastrado no banco")


# --------------- Arquivo de Estoque ---------------
def readStockFile(file):
    for value in file.iter_rows(min_row=4, min_col=3, max_col=6, values_only=True):
        updateStockAndSave(value[0], value[1], id, value[3], value[2])


def updateStockAndSave(insumo, quantidade, id, alerta, preco=0):
    connection = create_connection(
        "us-cdbr-east-06.cleardb.net", "b090112be85288", "2f84fdce", "heroku_7324e25c80c3d90")
    cursor = connection.cursor()
    if (preco != 0):
        # Entrará aqui se o arquivo for para atualizacao de estoque.
        stmt = (
            "INSERT INTO stock (product, amount, price, warning) VALUES (%s, %s , %s, %s)")
        values = (insumo, quantidade, preco, alerta)
        cursor.execute(stmt, values)
        connection.commit()
        cursor.close
        print(
            f"Foi adicionado {quantidade} unidades do insumo {insumo} ajustado ao custo base de {preco} - alerta de baixo estoque settado em {alerta} unidades")
    else:
        # Entrará aqui se o arquivo for para a retirada de estoque.
        stmt = ("UPDATE stock SET amount = amount - %s WHERE product = %s")
        data = (quantidade, insumo)
        cursor.execute(stmt, data)
        connection.commit()
        cursor.close
        print(
            f"Foi retirado {quantidade} unidades do insumo {insumo} no estoque")


# --------------- Rotina de Faturamento ---------------
def createBilling():
    connection = create_connection(
        "us-cdbr-east-06.cleardb.net", "b090112be85288", "2f84fdce", "heroku_7324e25c80c3d90")
    cursor = connection.cursor()

    # Vai checar quais os clientes finais possuem faturamentos no dia
    print("Gerando Faturamento")
    stmt = ("SELECT CURDATE()")
    cursor.execute(stmt)

    for row in cursor:
        date = row[0]

    day = date.strftime("%d")
    stmt = ("SELECT userPix,cons.clientName,email,city,userName,GROUP_CONCAT(price),GROUP_CONCAT(product),GROUP_CONCAT(amount) FROM data,cons WHERE billingDay = %s and data.clientName = cons.clientName group by userPix")
    d = (day,)
    cursor.execute(stmt, d)
    row = cursor.fetchall()

    for all in row:
        # print(all[0], all[1], all[2], all[3], all[4], all[5], all[6], all[7])
        priceList = all[5].split(',')
        total = sum(float(i) for i in priceList)
        total = int(total * 100)

        currentUser = User(all[4], "null", all[0])
        currentClient = Client(all[1], all[2], all[3], "null")
        currentProduct = Product(all[6].split(
            ','), all[7].split(','), priceList)

        pix = generateQRCode(all[0], all[3], all[4], str(total), all[1])

        qr = f"/content/static/generatedpix/pix-{all[1]}.png"

        sendBillingMail(qr, currentClient, currentUser, currentProduct, total)

    print("Fechando Faturamento")


def generateQRCode(userPix, city, username, value, clientName):
    pix = PixQrCode(username, userPix, city, value)

    if pix.is_valid():
        pix.save_qrcode(
            "/content/static/generatedpix", f"pix-{clientName}")
        return pix.generate_code()
    else:
        return "Dados para Pix inválidos."


# --------------- Envio de Email ---------------

def sendStockAlertMail(product, user):
    app_email = 'labprojetospuc@gmail.com'
    app_password = 'npfyvwfaljvlvplv'  # Token for gmail
    to = user.email

    productList = product.name
    amountList = product.amount
    priceList = product.unitPrice

    list = ""

    for num in range(len(productList)):
        lista = list + \
            f'<span style="border-bottom: 1px solid lightgray;display: inline-block;width: 100%;"><p style="text-align: left;display: inline-block;width: 70%;">{productList[num]}</p><i>{amountList[num]}</i><a style="text-align: right;display: inline-block;width: 15%;position: relative;">R${priceList[num]}</a></span>'

    subject = 'Alerta de produtos esgotando do estoque'
    content = ['<div style="text-align: center;background-color: #e1e1e1;"><div>',
               f'<h2>Olá {user.name},</h2><p style="font-size: large">Este é um email de aviso referente aos produtos que estão acabando de seu estoque. <br> Para adicionar mais no sistema, utilize a mesma template de estoque APENAS com os produtos que irá acrescentar a ele.</p></div>',
               f'<div style="border: solid 1px gray; width: 1000px; margin: 0 auto; background-color: white; box-shadow: 2px 2px #86868621"><header>',
               yagmail.inline("/content/static/img/title.png"),
               f'</header><section><article><h1>Estoque abaixo do limite</h1><div style="width: 100%;"><a style="width: 65%; padding-left: 3%;display: inline-block;position: relative;font-weight: bold;text-align: left;">Produto</a><a style="width: 12.5%; padding-left: 2.5%;display: inline-block;position: relative;font-weight: bold;text-align: left;">Quantidade</a><a style="width: 9%; padding-left: 1%;display: inline-block;position: relative;font-weight: bold;text-align: left;">Preço(un.)</a></div>',
               lista,
               f'</article></section></div><footer style="margin-top: 10px;color: rgb(143, 143, 143);font-family: sans-serif;"><p style="padding: 0px;margin: 5px;">Ward - Automated Billing Service</p><p style="padding: 0px;margin: 5px;">Av. Padre Cletus Francis Cox, 1661 - Country Club, Poços de Caldas - MG, 37714-620</p><p style="padding: 0px;margin: 5px;">Nos contate em labprojetospuc@gmail.com</p></footer></div>']

    with yagmail.SMTP(app_email, app_password) as yag:
        yag.send(to, subject, content)
        print('Sent email successfully')


def sendBillingMail(qr, client, user, product, total):
    userEmail = 'labprojetospuc@gmail.com'
    app_password = 'npfyvwfaljvlvplv'  # Token for gmail
    to = client.email

    productList = product.name
    amountList = product.amount
    priceList = product.unitPrice
    total = str(total)[:-2]
    virgula = ',00'
    total = total + virgula

    list = ""
    lista = []

    for num in range(len(productList)):
        list = f'<span style="border-bottom: 1px solid lightgray;display: inline-block;width: 100%;"><p style="text-align: left;display: inline-block;width: 70%;">{productList[num]}</p><i>{amountList[num]}</i><a style="text-align: right;display: inline-block;width: 15%;position: relative;">R${priceList[num]}</a></span>'
        lista.append(list)

    items = ["&emsp;{}".format(s) for s in lista]
    items = "".join(items)
    subject = f'Cobrança à {user.name}'
    content = ['<div style="text-align: center;background-color: #e1e1e1;"><div>',
               f'<h2>Olá {client.name},</h2><p style="font-size: large">Este é um email referente à cobrança gerada por {user.name}. Em anexo segue o código pix referente ao valor de R${total}.</p></div>',
               f'<div style="border: solid 1px gray; width: 1000px; margin: 0 auto; background-color: white; box-shadow: 2px 2px #86868621"><header>',
               yagmail.inline("/content/static/img/title.png"),
               f'</header><section><article><h1>Resumo da compra</h1><div style="width: 100%;"><a style="width: 65%; padding-left: 3%;display: inline-block;position: relative;font-weight: bold;text-align: left;">Produto</a><a style="width: 12.5%; padding-left: 2.5%;display: inline-block;position: relative;font-weight: bold;text-align: left;">Quantidade</a><a style="width: 9%; padding-left: 1%;display: inline-block;position: relative;font-weight: bold;text-align: left;">Preço</a></div>',
               items,
               f'<div style="border-top: 2px solid black"><p style="width: 95%;text-align: right;font-family: sans-serif;font-size: large;font-weight: bold;">Total: R${total}</p></div></article></section></div>',
               yagmail.inline(
                   f"/content/static/generatedpix/pix-{client.name}.png"),
               '<footer style="margin-top: 10px;color: rgb(143, 143, 143);font-family: sans-serif;"><p style="padding: 0px;margin: 5px;">Ward - Automated Billing Service</p><p style="padding: 0px;margin: 5px;">Av. Padre Cletus Francis Cox, 1661 - Country Club, Poços de Caldas - MG, 37714-620</p><p style="padding: 0px;margin: 5px;">Nos contate em labprojetospuc@gmail.com</p></footer></div>']
    attachments = [qr]

    with yagmail.SMTP(userEmail, app_password) as yag:
        yag.send(to, subject, content, attachments)
        print('Sent email successfully')


# --------------- Rotina de Controle de Estoque ---------------
def checkStock():
    # Vai checar, para cada user, como estao os estoques
    connection = create_connection(
        "us-cdbr-east-06.cleardb.net", "b090112be85288", "2f84fdce", "heroku_7324e25c80c3d90")
    cursor = connection.cursor()
    stmt = ("SELECT GROUP_CONCAT(price),GROUP_CONCAT(product),GROUP_CONCAT(amount),userEmail,userName FROM stock,data WHERE amount = 0 group by product")
    cursor.execute(stmt)
    row = cursor.fetchall()
    for all in row:
        currentUser = User(all[4], all[3], "pix")
        currentProduct = Product(all[1].split(
            ','), all[2].split(','), all[0].split(','))
        sendStockAlertMail(currentProduct, currentUser)

    stmt = ("SELECT GROUP_CONCAT(price),GROUP_CONCAT(product),GROUP_CONCAT(amount),userEmail,userName FROM stock,data WHERE amount = warning GROUP BY product,amount,userEmail")
    cursor.execute(stmt)
    row = cursor.fetchall()
    for all in row:
        currentUser = User(all[4], all[3], "pix")
        currentProduct = Product(all[1].split(
            ','), all[2].split(','), all[0].split(','))
        sendStockAlertMail(currentProduct, currentUser)
    cursor.close

# --------------- Controle de Rotinas ---------------


def callDailyRoutines():
    try:
        createBilling()
    except ValueError:
        print("Erro ao gerar faturamento!")
    try:
        checkStock()
    except ValueError:
        print("Erro ao realizar a verificação do estoque!")


# app = Flask(__name__, template_folder='/content/templates')
# app._static_folder = '/content'
# run_with_ngrok(app)


# @app.route("/")
# def home():
#     return render_template('index.html')


# app.run()


# form = cgi.FieldStorage()
# fileitem = form['filename']
# code = sheet["AA1"].value

# if code == 'expense':
#     readExpenseFile(sheet)
# elif code == 'stock':
#     readStockFile(sheet)
# elif code == 'dados':
#     readDataFile(sheet)
# else:
#     print('O arquivo não pode ser validado. Por favor utilize os templates disponíveis no github!')

try:
    # Aqui entrar o arquivo de upload
    input_sheet_cons = load_workbook(
        filename="/content/excel/Exemplo-Consumo.xlsx")
    input_sheet_data = load_workbook(
        filename="/content/excel/Exemplo-Dados.xlsx")
    input_sheet_stock = load_workbook(
        filename="/content/excel/Exemplo-Estoque.xlsx")
    sheet = input_sheet_data.active
    readDataFile(sheet)
    #sheet = input_sheet_stock.active
    # readStockFile(sheet)
    sheet = input_sheet_cons.active
    readExpenseFile(sheet)
except ValueError:
    print("Erro no upload de arquivo. Será aceita apenas a extensão .xlsx")

callDailyRoutines()
